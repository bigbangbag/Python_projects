from openpyxl import Workbook

# Create a new workbook
wb = Workbook()

# Get the active worksheet
ws = wb.active

# Rename the worksheet
ws.title = "MySheet"

# Add some data
ws["A1"] = "Name"
ws["B1"] = "Age"
ws["C1"] = "City"

data = [
    ["Alice", 30, "New York"],
    ["Bob", 25, "Los Angeles"],
    ["Charlie", 35, "Chicago"]
]

for row in data:
    ws.append(row)

# Save the workbook to a file
wb.save("sample.xlsx")

print("Excel file 'sample.xlsx' has been created successfully!")
